from __future__ import annotations

import calendar
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import (
    AttendanceCode,
    AttendanceEntry,
    Employee,
    EmployeeStatus,
    Holiday,
    LeaveRequest,
    LeaveRequestStatus,
    LeaveType,
)

MONTHS_FR = [
    "",
    "Janvier",
    "Février",
    "Mars",
    "Avril",
    "Mai",
    "Juin",
    "Juillet",
    "Août",
    "Septembre",
    "Octobre",
    "Novembre",
    "Décembre",
]

HOLIDAY_CODE = "F"


def _leave_code(leave_type: LeaveType) -> str:
    if leave_type.code_court:
        return leave_type.code_court
    return leave_type.libelle[:3].upper()


def build_monthly_state(db: Session, mois: int, annee: int) -> dict:
    """Merges, per active employee per day of the month: pointeuse import
    (AttendanceEntry), approved congé (LeaveRequest — takes precedence for
    payroll purposes), and holidays. Does NOT include variable pay
    (avances/primes/commissions) or disciplinary suspensions — those modules
    don't exist yet in this rebuild; see README known gaps. A day carrying
    both a pointeuse entry and an approved leave is flagged as a conflict
    rather than silently dropping one."""
    nb_jours = calendar.monthrange(annee, mois)[1]
    start, end = date(annee, mois, 1), date(annee, mois, nb_jours)

    employees = (
        db.query(Employee)
        .join(EmployeeStatus, Employee.status_id == EmployeeStatus.id, isouter=True)
        .filter(EmployeeStatus.is_active_status.is_(True))
        .order_by(Employee.nom, Employee.prenom)
        .all()
    )

    entries = (
        db.query(AttendanceEntry)
        .filter(AttendanceEntry.date >= start, AttendanceEntry.date <= end)
        .all()
    )
    entries_by_key = {(e.employee_id, e.date): e for e in entries}

    leave_requests = (
        db.query(LeaveRequest)
        .filter(
            LeaveRequest.status == LeaveRequestStatus.APPROVED,
            LeaveRequest.date_debut <= end,
            LeaveRequest.date_fin >= start,
        )
        .all()
    )
    leaves_by_employee: dict[int, list[LeaveRequest]] = {}
    for req in leave_requests:
        leaves_by_employee.setdefault(req.employee_id, []).append(req)

    holidays = {
        h.date for h in db.query(Holiday).filter(Holiday.date >= start, Holiday.date <= end).all()
    }

    rows = []
    nb_conflits = 0
    for employee in employees:
        emp_leaves = leaves_by_employee.get(employee.id, [])
        days = []
        for day_num in range(1, nb_jours + 1):
            current = date(annee, mois, day_num)
            entry = entries_by_key.get((employee.id, current))
            leave = next((r for r in emp_leaves if r.date_debut <= current <= r.date_fin), None)

            code: str | None = None
            conflict = False
            if leave is not None:
                code = _leave_code(leave.leave_type)
                if entry is not None and entry.code_id is not None:
                    conflict = True
                    nb_conflits += 1
            elif entry is not None and entry.code is not None:
                code = entry.code.code_court
            elif current in holidays:
                code = HOLIDAY_CODE

            days.append({"date": current.isoformat(), "code": code, "conflict": conflict})

        rows.append(
            {
                "employee_id": employee.id,
                "nom_complet": employee.full_name,
                "departement": employee.department.nom if employee.department else None,
                "days": days,
            }
        )

    return {
        "mois": mois,
        "annee": annee,
        "nb_jours": nb_jours,
        "rows": rows,
        "nb_conflits": nb_conflits,
    }


def export_monthly_state_xlsx(db: Session, mois: int, annee: int) -> bytes:
    state = build_monthly_state(db, mois, annee)
    nb_jours = state["nb_jours"]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "État de présence"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="E6EEF9", end_color="E6EEF9", fill_type="solid")
    conflict_fill = PatternFill(start_color="FDEAE3", end_color="FDEAE3", fill_type="solid")

    title_cell = sheet.cell(row=1, column=1, value=f"État de présence — {MONTHS_FR[mois]} {annee}")
    title_cell.font = Font(bold=True, size=13)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_jours + 2)

    header_row = 3
    sheet.cell(row=header_row, column=1, value="Collaborateur").font = bold
    sheet.cell(row=header_row, column=2, value="Département").font = bold
    for day_num in range(1, nb_jours + 1):
        cell = sheet.cell(row=header_row, column=2 + day_num, value=day_num)
        cell.font = bold
        cell.fill = header_fill

    row_idx = header_row + 1
    for row in state["rows"]:
        sheet.cell(row=row_idx, column=1, value=row["nom_complet"])
        sheet.cell(row=row_idx, column=2, value=row["departement"] or "—")
        for i, day in enumerate(row["days"]):
            cell = sheet.cell(row=row_idx, column=3 + i, value=day["code"] or "")
            if day["conflict"]:
                cell.fill = conflict_fill
        row_idx += 1

    sheet.column_dimensions[get_column_letter(1)].width = 24
    sheet.column_dimensions[get_column_letter(2)].width = 18
    for col_idx in range(3, nb_jours + 3):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 4

    legend = workbook.create_sheet("Légende")
    legend.cell(row=1, column=1, value="Code").font = bold
    legend.cell(row=1, column=2, value="Libellé").font = bold
    legend_row = 2
    for code in (
        db.query(AttendanceCode)
        .filter(AttendanceCode.is_active.is_(True))
        .order_by(AttendanceCode.code_court)
        .all()
    ):
        legend.cell(row=legend_row, column=1, value=code.code_court)
        legend.cell(row=legend_row, column=2, value=code.libelle)
        legend_row += 1
    for leave_type in (
        db.query(LeaveType).filter(LeaveType.is_active.is_(True)).order_by(LeaveType.libelle).all()
    ):
        legend.cell(row=legend_row, column=1, value=_leave_code(leave_type))
        legend.cell(row=legend_row, column=2, value=f"{leave_type.libelle} (congé)")
        legend_row += 1
    legend.cell(row=legend_row, column=1, value=HOLIDAY_CODE)
    legend.cell(row=legend_row, column=2, value="Jour férié")
    legend.column_dimensions["A"].width = 10
    legend.column_dimensions["B"].width = 30

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
